from datetime import datetime, timezone
from io import BytesIO

from aiogram import Bot, F, Router
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import BaseFilter, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import config
import db
import texts

PAGE_SIZE = 8

# Ввод в FSM-шагах: любой текст, кроме команд. Иначе /start и прочие
# команды попали бы в название даты вместо того, чтобы сработать.
PLAIN_TEXT = F.text & ~F.text.startswith("/")

STATUS_ICON = {db.PENDING: "⏳", db.CONFIRMED: "✅", db.CANCELLED: "❌"}
STATUS_TEXT = {
    db.PENDING: "ждёт подтверждения",
    db.CONFIRMED: "подтверждена",
    db.CANCELLED: "отменена",
}

OUTCOME_ICON = {db.PASSED: "🎓", db.FAILED: "📕", db.NO_SHOW: "🚷"}
OUTCOME_TEXT = {
    db.PASSED: "явился и сдал",
    db.FAILED: "явился, не сдал",
    db.NO_SHOW: "не пришёл",
}


class IsAdmin(BaseFilter):
    async def __call__(self, event: Message | CallbackQuery) -> bool:
        return event.from_user is not None and event.from_user.id in config.ADMIN_IDS


router = Router(name="admin")
router.message.filter(IsAdmin())
router.callback_query.filter(IsAdmin())


class AdminFlow(StatesGroup):
    date_title = State()
    date_limit_new = State()
    date_limit_edit = State()
    date_title_edit = State()
    date_value = State()
    date_value_edit = State()
    dates_bulk = State()


# ---------- ХЕЛПЕРЫ ----------

def fmt_dt(raw: str | None) -> str:
    """created_at лежит в базе в UTC, показываем в ташкентском времени.

    Заявки, созданные до перехода на timezone-aware время, записаны через
    utcnow() и лежат без смещения — их тоже считаем UTC.
    """
    if not raw:
        return "—"
    try:
        dt = datetime.fromisoformat(raw)
    except ValueError:
        return raw[:19].replace("T", " ")
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(config.TZ).strftime("%d.%m.%Y %H:%M")


def fmt_limit(seats_limit: int) -> str:
    return str(seats_limit) if seats_limit else "без лимита"


def back_button(callback_data: str, text: str = "⬅️ Назад") -> InlineKeyboardButton:
    return InlineKeyboardButton(text=text, callback_data=callback_data)


async def show(callback: CallbackQuery, text: str, keyboard: InlineKeyboardMarkup):
    """Перерисовывает экран админки на месте; если нельзя — шлёт новым сообщением."""
    try:
        await callback.message.edit_text(text, reply_markup=keyboard)
    except TelegramBadRequest:
        await callback.message.answer(text, reply_markup=keyboard)


def menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📅 Даты", callback_data="a:dates")],
        [InlineKeyboardButton(text="👥 Записавшиеся", callback_data="a:bdates")],
        [InlineKeyboardButton(text="📊 Статистика", callback_data="a:stats")],
        [InlineKeyboardButton(text="🔻 Воронка", callback_data="a:fn:today")],
        [InlineKeyboardButton(text="📥 Экспорт в Excel", callback_data="a:export")],
    ])


MENU_TEXT = "🛠 Админ-панель\n\nВыберите раздел:"


# ---------- ВХОД ----------

@router.message(Command("admin"))
async def admin_entry(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(MENU_TEXT, reply_markup=menu_keyboard())


@router.message(Command("cancel"))
async def admin_cancel(message: Message, state: FSMContext):
    if await state.get_state() is None:
        return
    await state.clear()
    await message.answer("Отменено.", reply_markup=menu_keyboard())


@router.callback_query(F.data == "a:menu")
async def admin_menu(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await show(callback, MENU_TEXT, menu_keyboard())
    await callback.answer()


# ---------- ДАТЫ ----------

def is_past(exam_date: str | None) -> bool:
    return bool(exam_date) and exam_date < db.today().isoformat()


def fmt_exam_date(exam_date: str | None) -> str:
    if not exam_date:
        return "не задана"
    return datetime.strptime(exam_date, "%Y-%m-%d").strftime("%d.%m.%Y")


def date_button(row: tuple) -> list[InlineKeyboardButton]:
    date_id, title, seats_limit, is_active, confirmed, pending, exam_date = row
    taken = confirmed + pending
    seats = f"{taken}/{seats_limit}" if seats_limit else f"{taken}"
    if not is_active:
        mark = "🚫 "
    elif is_past(exam_date):
        mark = "⌛ "
    else:
        mark = ""
    return [InlineKeyboardButton(
        text=f"{mark}{title} — {seats}", callback_data=f"a:d:{date_id}"
    )]


async def render_dates(callback: CallbackQuery):
    """Только активные даты. Скрытые убраны с глаз, но не потеряны:
    они за кнопкой «Скрытые», откуда их можно вернуть в запись."""
    dates = await db.get_dates_overview()
    active = [d for d in dates if d[3]]
    hidden = [d for d in dates if not d[3]]

    rows = [date_button(d) for d in active]
    if hidden:
        rows.append([InlineKeyboardButton(
            text=f"🚫 Скрытые ({len(hidden)})", callback_data="a:dhid"
        )])
    rows.append([InlineKeyboardButton(text="➕ Добавить дату", callback_data="a:dadd")])
    rows.append([InlineKeyboardButton(text="📋 Добавить списком", callback_data="a:dbulk")])
    rows.append([back_button("a:menu", "⬅️ В меню")])

    if active:
        text = ("📅 Даты экзамена\n\n"
                "В кнопках — занято мест (записи в ожидании + подтверждённые).")
    elif hidden:
        text = "📅 Даты экзамена\n\nАктивных дат нет — все скрыты из записи."
    else:
        text = "📅 Даты экзамена\n\nПока ни одной даты нет."

    await show(callback, text, InlineKeyboardMarkup(inline_keyboard=rows))


async def render_hidden_dates(callback: CallbackQuery):
    hidden = [d for d in await db.get_dates_overview() if not d[3]]
    if not hidden:
        await render_dates(callback)
        return

    rows = [date_button(d) for d in hidden]
    rows.append([back_button("a:dates")])
    await show(
        callback,
        "🚫 Скрытые даты\n\n"
        "Клиенты их не видят. Откройте дату, чтобы вернуть её в запись "
        "или удалить окончательно.",
        InlineKeyboardMarkup(inline_keyboard=rows),
    )


@router.callback_query(F.data == "a:dhid")
async def hidden_dates_list(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await render_hidden_dates(callback)
    await callback.answer()



@router.callback_query(F.data == "a:dates")
async def dates_list(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await render_dates(callback)
    await callback.answer()


async def render_date_detail(callback: CallbackQuery, date_id: int) -> bool:
    row = await db.get_date_overview(date_id)
    if not row:
        await render_dates(callback)
        return False

    _, title, seats_limit, is_active, confirmed, pending, exam_date = row
    taken = confirmed + pending
    free = "—" if not seats_limit else max(seats_limit - taken, 0)

    if not is_active:
        status = "скрыта из записи"
    elif is_past(exam_date):
        status = "прошла — клиентам не показывается"
    else:
        status = "активна"

    text = (
        f"📅 {title}\n\n"
        f"Дата: {fmt_exam_date(exam_date)}\n"
        f"Статус: {status}\n"
        f"Лимит мест: {fmt_limit(seats_limit)}\n"
        f"Свободно: {free}\n\n"
        f"✅ Подтверждено: {confirmed}\n"
        f"⏳ Ждут подтверждения: {pending}"
    )

    rows = [
        [InlineKeyboardButton(
            text="👥 Записавшиеся", callback_data=f"a:bk:{date_id}:0"
        )],
        [InlineKeyboardButton(
            text="📆 Изменить дату", callback_data=f"a:ddate:{date_id}"
        )],
        [InlineKeyboardButton(
            text="🔢 Изменить лимит", callback_data=f"a:dlim:{date_id}"
        )],
        [InlineKeyboardButton(
            text="✏️ Переименовать", callback_data=f"a:dren:{date_id}"
        )],
    ]
    if is_active:
        rows.append([InlineKeyboardButton(
            text="🗑 Удалить дату", callback_data=f"a:ddel:{date_id}"
        )])
    else:
        rows.append([InlineKeyboardButton(
            text="♻️ Вернуть в запись", callback_data=f"a:drst:{date_id}"
        )])
    rows.append([back_button("a:dates" if is_active else "a:dhid")])

    await show(callback, text, InlineKeyboardMarkup(inline_keyboard=rows))
    return True


@router.callback_query(F.data.startswith("a:d:"))
async def date_detail(callback: CallbackQuery):
    date_id = int(callback.data.split(":")[2])
    if await render_date_detail(callback, date_id):
        await callback.answer()
    else:
        await callback.answer("Дата не найдена.", show_alert=True)


@router.callback_query(F.data == "a:dadd")
async def date_add_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AdminFlow.date_value)
    await show(
        callback,
        "➕ Новая дата\n\n"
        "Пришлите дату экзамена.\n"
        "Формат: 07.09.2026 (можно 7.9.26 или 07.09 — год текущий).\nНесуществующие числа вроде 31.09 бот не примет.\n\n"
        "/cancel — отмена",
        InlineKeyboardMarkup(inline_keyboard=[[back_button("a:dates", "⬅️ Отмена")]]),
    )
    await callback.answer()


@router.message(AdminFlow.date_value, PLAIN_TEXT)
async def date_add_value(message: Message, state: FSMContext):
    iso = db.parse_exam_date(message.text)
    if not iso:
        await message.answer(
            "Не понял дату, либо такого числа нет в календаре.\n\n"
            "Формат: 07.09.2026 (можно 7.9.26 или 07.09 — год текущий).\nНесуществующие числа вроде 31.09 бот не примет."
        )
        return

    # название выводим из даты: так оно не разъедется с сортировкой,
    # а переименовать под свой вкус можно потом
    title = fmt_exam_date(iso)
    await state.update_data(title=title, exam_date=iso)
    await state.set_state(AdminFlow.date_limit_new)
    await message.answer(
        f"Название: {title}\n\n"
        "Теперь пришлите лимит мест числом. 0 — без лимита.\n\n"
        "/cancel — отмена"
    )


@router.message(AdminFlow.date_limit_new, PLAIN_TEXT)
async def date_add_limit(message: Message, state: FSMContext):
    if not message.text.strip().isdigit():
        await message.answer("Нужно целое число (0 — без лимита).")
        return

    seats_limit = int(message.text.strip())
    data = await state.get_data()
    await state.clear()

    date_id = await db.add_date(data["title"], seats_limit, data.get("exam_date"))
    if date_id is None:
        await message.answer(
            f"Дата «{data['title']}» уже есть в списке.",
            reply_markup=menu_keyboard(),
        )
        return

    await message.answer(
        f"✅ Дата «{data['title']}» добавлена. Лимит мест: {fmt_limit(seats_limit)}.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📅 К списку дат", callback_data="a:dates")],
            [back_button("a:menu", "⬅️ В меню")],
        ]),
    )


@router.callback_query(F.data == "a:dbulk")
async def dates_bulk_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AdminFlow.dates_bulk)
    await show(
        callback,
        "📋 Добавить даты списком\n\n"
        "Пришлите одним сообщением, по одной дате в строке:\n"
        "07.09.2026 5\n"
        "09.09.2026 25\n"
        "11.09.2026 25\n\n"
        "Число после даты — лимит мест (0 или без числа — без лимита).\n"
        "Уже существующие даты и строки с ошибками бот пропустит и покажет какие.\n\n"
        "/cancel — отмена",
        InlineKeyboardMarkup(inline_keyboard=[[back_button("a:dates", "⬅️ Отмена")]]),
    )
    await callback.answer()


@router.message(AdminFlow.dates_bulk, PLAIN_TEXT)
async def dates_bulk_apply(message: Message, state: FSMContext):
    added, duplicate, bad = [], [], []
    for raw in message.text.splitlines():
        raw = raw.strip()
        if not raw:
            continue
        # разделителем может быть пробел или тире; тире внутри 2026-09-07 не трогаем
        parts = [p for p in raw.split() if p not in ("-", "—", "–")]
        iso = db.parse_exam_date(parts[0]) if parts else None
        if not iso:
            bad.append(raw)
            continue
        seats_limit = int(parts[-1]) if len(parts) > 1 and parts[-1].isdigit() else 0
        title = fmt_exam_date(iso)
        if await db.add_date(title, seats_limit, iso) is None:
            duplicate.append(title)
        else:
            added.append(f"{title} — {fmt_limit(seats_limit)}")

    await state.clear()

    report = []
    if added:
        report.append("✅ Добавлено ({}):\n".format(len(added)) + "\n".join(added))
    if duplicate:
        report.append("↩️ Уже были ({}):\n".format(len(duplicate)) + "\n".join(duplicate))
    if bad:
        report.append("⚠️ Не разобрал ({}):\n".format(len(bad)) + "\n".join(bad))
    if not report:
        report.append("Ни одной строки с датой не нашёл.")

    await message.answer(
        "\n\n".join(report),
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📅 К списку дат", callback_data="a:dates")],
            [back_button("a:menu", "⬅️ В меню")],
        ]),
    )



@router.callback_query(F.data.startswith("a:ddate:"))
async def date_value_start(callback: CallbackQuery, state: FSMContext):
    date_id = int(callback.data.split(":")[2])
    row = await db.get_date_overview(date_id)
    if not row:
        await callback.answer("Дата не найдена.", show_alert=True)
        return

    await state.set_state(AdminFlow.date_value_edit)
    await state.update_data(date_id=date_id)
    await show(
        callback,
        f"📆 Дата для «{row[1]}»\n\n"
        f"Сейчас: {fmt_exam_date(row[6])}\n\n"
        "Пришлите новую дату. От неё зависят порядок в списке и то, "
        f"когда дата перестанет показываться клиентам.\n"
        "Формат: 07.09.2026 (можно 7.9.26 или 07.09 — год текущий).\nНесуществующие числа вроде 31.09 бот не примет.\n\n"
        "/cancel — отмена",
        InlineKeyboardMarkup(inline_keyboard=[
            [back_button(f"a:d:{date_id}", "⬅️ Отмена")]
        ]),
    )
    await callback.answer()


@router.message(AdminFlow.date_value_edit, PLAIN_TEXT)
async def date_value_save(message: Message, state: FSMContext):
    iso = db.parse_exam_date(message.text)
    if not iso:
        await message.answer(
            "Не понял дату, либо такого числа нет в календаре.\n\n"
            "Формат: 07.09.2026 (можно 7.9.26 или 07.09 — год текущий).\nНесуществующие числа вроде 31.09 бот не примет."
        )
        return

    data = await state.get_data()
    date_id = data["date_id"]
    await state.clear()

    if not await db.set_exam_date(date_id, iso):
        await message.answer("Дата не найдена.", reply_markup=menu_keyboard())
        return

    warning = ""
    if is_past(iso):
        warning = f"\n\n⚠️ Это прошедшее число — клиентам дата больше не покажется."

    await message.answer(
        f"✅ Дата экзамена: {fmt_exam_date(iso)}.{warning}",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📅 К дате", callback_data=f"a:d:{date_id}")],
            [back_button("a:dates", "⬅️ К списку дат")],
        ]),
    )


@router.callback_query(F.data.startswith("a:dlim:"))
async def date_limit_start(callback: CallbackQuery, state: FSMContext):
    date_id = int(callback.data.split(":")[2])
    row = await db.get_date_overview(date_id)
    if not row:
        await callback.answer("Дата не найдена.", show_alert=True)
        return

    _, title, seats_limit, _, confirmed, pending, _ = row
    await state.set_state(AdminFlow.date_limit_edit)
    await state.update_data(date_id=date_id)

    await show(
        callback,
        f"✏️ Лимит мест для «{title}»\n\n"
        f"Сейчас: {fmt_limit(seats_limit)}\n"
        f"Уже занято: {confirmed + pending}\n\n"
        "Пришлите новый лимит числом. 0 — без лимита.\n\n"
        "/cancel — отмена",
        InlineKeyboardMarkup(inline_keyboard=[
            [back_button(f"a:d:{date_id}", "⬅️ Отмена")]
        ]),
    )
    await callback.answer()


@router.message(AdminFlow.date_limit_edit, PLAIN_TEXT)
async def date_limit_save(message: Message, state: FSMContext):
    if not message.text.strip().isdigit():
        await message.answer("Нужно целое число (0 — без лимита).")
        return

    seats_limit = int(message.text.strip())
    data = await state.get_data()
    date_id = data["date_id"]
    await state.clear()

    await db.set_date_limit(date_id, seats_limit)
    row = await db.get_date_overview(date_id)
    title = row[1] if row else f"#{date_id}"

    await message.answer(
        f"✅ Лимит для «{title}» — {fmt_limit(seats_limit)}.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📅 К дате", callback_data=f"a:d:{date_id}")],
            [back_button("a:menu", "⬅️ В меню")],
        ]),
    )


@router.callback_query(F.data.startswith("a:dren:"))
async def date_rename_start(callback: CallbackQuery, state: FSMContext):
    date_id = int(callback.data.split(":")[2])
    row = await db.get_date_overview(date_id)
    if not row:
        await callback.answer("Дата не найдена.", show_alert=True)
        return

    title = row[1]
    await state.set_state(AdminFlow.date_title_edit)
    await state.update_data(date_id=date_id)

    await show(
        callback,
        f"✏️ Переименование даты «{title}»\n\n"
        "Пришлите новое название так, как его увидит клиент.\n"
        "Заявки и лимит останутся на месте.\n\n"
        "/cancel — отмена",
        InlineKeyboardMarkup(inline_keyboard=[
            [back_button(f"a:d:{date_id}", "⬅️ Отмена")]
        ]),
    )
    await callback.answer()


@router.message(AdminFlow.date_title_edit, PLAIN_TEXT)
async def date_rename_save(message: Message, state: FSMContext):
    title = message.text.strip()
    if not title:
        await message.answer("Название не может быть пустым. Пришлите другое.")
        return

    data = await state.get_data()
    date_id = data["date_id"]
    result = await db.rename_date(date_id, title)

    if result == "duplicate":
        # состояние не сбрасываем — админ может сразу прислать другое название
        await message.answer(f"Дата «{title}» уже есть в списке. Пришлите другое название.")
        return

    await state.clear()

    if result == "missing":
        await message.answer("Дата не найдена — похоже, её успели удалить.",
                             reply_markup=menu_keyboard())
        return

    await message.answer(
        f"✅ Новое название: «{title}».",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📅 К дате", callback_data=f"a:d:{date_id}")],
            [back_button("a:menu", "⬅️ В меню")],
        ]),
    )


@router.callback_query(F.data.startswith("a:ddel:"))
async def date_delete_confirm(callback: CallbackQuery):
    date_id = int(callback.data.split(":")[2])
    row = await db.get_date_overview(date_id)
    if not row:
        await callback.answer("Дата не найдена.", show_alert=True)
        return

    _, title, _, _, confirmed, pending, _ = row
    total = confirmed + pending

    text = f"🗑 Удалить дату «{title}»?"
    if total:
        text += (
            f"\n\nПо ней уже есть заявки ({total}). Дата будет скрыта из записи, "
            "а заявки и история останутся."
        )

    await show(callback, text, InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🗑 Да, удалить", callback_data=f"a:ddel1:{date_id}")],
        [back_button(f"a:d:{date_id}", "⬅️ Отмена")],
    ]))
    await callback.answer()


@router.callback_query(F.data.startswith("a:ddel1:"))
async def date_delete_do(callback: CallbackQuery):
    date_id = int(callback.data.split(":")[2])
    result = await db.delete_date(date_id)
    await callback.answer("Удалено" if result == "deleted" else "Дата скрыта из записи")
    await render_dates(callback)


@router.callback_query(F.data.startswith("a:drst:"))
async def date_restore(callback: CallbackQuery):
    date_id = int(callback.data.split(":")[2])
    await db.restore_date(date_id)
    await callback.answer("Дата снова доступна для записи")
    await render_date_detail(callback, date_id)


# ---------- ЗАПИСАВШИЕСЯ ----------

@router.callback_query(F.data == "a:bdates")
async def bookings_dates(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    dates = await db.get_dates_overview()

    rows = [
        [InlineKeyboardButton(
            text=f"{title} — {confirmed + pending}", callback_data=f"a:bk:{date_id}:0"
        )]
        for date_id, title, _, _, confirmed, pending, _ in dates
    ]
    rows.append([back_button("a:menu", "⬅️ В меню")])

    text = "👥 Записавшиеся\n\nВыберите дату:"
    if not dates:
        text = "👥 Записавшиеся\n\nДат пока нет."

    await show(callback, text, InlineKeyboardMarkup(inline_keyboard=rows))
    await callback.answer()


@router.callback_query(F.data.startswith("a:bk:"))
async def bookings_list(callback: CallbackQuery):
    _, _, raw_date_id, raw_page = callback.data.split(":")
    date_id, page = int(raw_date_id), int(raw_page)

    date_row = await db.get_date_overview(date_id)
    title = date_row[1] if date_row else f"#{date_id}"
    bookings = await db.get_bookings_for_date(date_id)

    if not bookings:
        await show(callback, f"👥 {title}\n\nЗаписей пока нет.", InlineKeyboardMarkup(
            inline_keyboard=[[back_button(f"a:d:{date_id}")]]
        ))
        await callback.answer()
        return

    pages = (len(bookings) - 1) // PAGE_SIZE + 1
    page = max(0, min(page, pages - 1))
    chunk = bookings[page * PAGE_SIZE:(page + 1) * PAGE_SIZE]

    rows = []
    for (booking_id, full_name, username, _, status, _, _, outcome, applicant_name) in chunk:
        icon = STATUS_ICON.get(status, "•")
        # если записывали друга, в кнопке важнее его имя, а не владельца аккаунта
        name = applicant_name or full_name or (f"@{username}" if username else "без имени")
        # у прошедших экзаменов сразу видно, кому итог ещё не проставили
        mark = OUTCOME_ICON.get(outcome, "") if outcome else ""
        name = name[:24]
        rows.append([InlineKeyboardButton(
            text=f"{icon}{mark} №{booking_id} · клиент: {name}",
            callback_data=f"a:b:{booking_id}",
        )])

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton(
            text="⬅️", callback_data=f"a:bk:{date_id}:{page - 1}"
        ))
    if page < pages - 1:
        nav.append(InlineKeyboardButton(
            text="➡️", callback_data=f"a:bk:{date_id}:{page + 1}"
        ))
    if nav:
        rows.append(nav)
    rows.append([back_button(f"a:d:{date_id}")])

    text = (
        f"👥 {title}\n"
        f"Всего записей: {len(bookings)}"
        + (f"   ·   стр. {page + 1}/{pages}" if pages > 1 else "")
        + "\n\n✅ подтверждена   ⏳ ждёт   ❌ отменена"
        + "\n🎓 сдал   📕 не сдал   🚷 не пришёл"
    )

    await show(callback, text, InlineKeyboardMarkup(inline_keyboard=rows))
    await callback.answer()


async def render_booking_detail(callback: CallbackQuery, booking_id: int) -> bool:
    row = await db.get_booking(booking_id)
    if not row:
        return False

    (_, user_id, full_name, username, date_id, status,
     confirmed_by_name, created_at, date_title,
     passport_file_id, receipt_file_id, cancelled_by_name,
     exam_date, outcome, outcome_by_name, outcome_at, applicant_name) = row

    lines = [
        f"Заявка №{booking_id}\n",
        f"Клиент: {full_name or '—'}"
        f"{' (@' + username + ')' if username else ''}",
        f"ID: {user_id}",
        *([f"Записан: {applicant_name}"] if applicant_name else []),
        f"Дата экзамена: {date_title or '—'}",
        f"Статус: {STATUS_ICON.get(status, '•')} {STATUS_TEXT.get(status, status)}",
        f"Подтвердил: {confirmed_by_name or '—'}",
    ]
    if status == db.CANCELLED:
        lines.append(f"Отменил: {cancelled_by_name or '—'}")
    if outcome:
        lines.append(
            f"Итог: {OUTCOME_ICON.get(outcome, '•')} {OUTCOME_TEXT.get(outcome, outcome)}"
            f" ({outcome_by_name or '—'}, {fmt_dt(outcome_at)})"
        )
    lines.append(f"Создана: {fmt_dt(created_at)}")
    text = "\n".join(lines)

    rows = []
    if passport_file_id or receipt_file_id:
        rows.append([InlineKeyboardButton(
            text="📎 Документы", callback_data=f"a:bdoc:{booking_id}"
        )])

    # итог проставляется только по подтверждённым заявкам и только после экзамена
    if status == db.CONFIRMED and is_past(exam_date):
        rows.append([InlineKeyboardButton(
            text="🎓 Явился и сдал", callback_data=f"a:bo:{booking_id}:{db.PASSED}"
        )])
        rows.append([InlineKeyboardButton(
            text="📕 Явился, не сдал", callback_data=f"a:bo:{booking_id}:{db.FAILED}"
        )])
        rows.append([InlineKeyboardButton(
            text="🚷 Не пришёл", callback_data=f"a:bo:{booking_id}:{db.NO_SHOW}"
        )])

    if status != db.CANCELLED:
        rows.append([InlineKeyboardButton(
            text="❌ Отменить запись", callback_data=f"a:bc:{booking_id}"
        )])
    rows.append([back_button(f"a:bk:{date_id}:0", "⬅️ К списку")])

    await show(callback, text, InlineKeyboardMarkup(inline_keyboard=rows))
    return True


@router.callback_query(F.data.startswith("a:b:"))
async def booking_detail(callback: CallbackQuery):
    booking_id = int(callback.data.split(":")[2])
    if await render_booking_detail(callback, booking_id):
        await callback.answer()
    else:
        await callback.answer("Заявка не найдена.", show_alert=True)


@router.callback_query(F.data.startswith("a:bdoc:"))
async def booking_documents(callback: CallbackQuery, bot: Bot):
    """Присылает админу паспорт и чек — file_id хранятся с момента подачи."""
    booking_id = int(callback.data.split(":")[2])
    row = await db.get_booking(booking_id)
    if not row:
        await callback.answer("Заявка не найдена.", show_alert=True)
        return

    sent = 0
    for file_id, label in ((row[9], "паспорт"), (row[10], "чек")):
        if not file_id:
            continue
        caption = f"Заявка №{booking_id}: {label}"
        try:
            await bot.send_photo(callback.from_user.id, file_id, caption=caption)
        except TelegramBadRequest:
            # клиент мог прислать документом, а не фото
            try:
                await bot.send_document(callback.from_user.id, file_id, caption=caption)
            except TelegramBadRequest as e:
                await callback.message.answer(f"⚠️ {caption} — не открылся: {e}")
                continue
        sent += 1

    await callback.answer("Отправил" if sent else "Документов нет")



@router.callback_query(F.data.startswith("a:bo:"))
async def booking_outcome(callback: CallbackQuery):
    _, _, raw_id, outcome = callback.data.split(":")
    booking_id = int(raw_id)

    row = await db.get_booking(booking_id)
    if not row:
        await callback.answer("Заявка не найдена.", show_alert=True)
        return
    if not is_past(row[12]):
        await callback.answer("Экзамен ещё не прошёл.", show_alert=True)
        return

    ok = await db.set_outcome(
        booking_id, outcome, callback.from_user.id, callback.from_user.full_name
    )
    if not ok:
        await callback.answer("Итог можно ставить только подтверждённым.", show_alert=True)
        return

    await db.log_event(row[1], db.STEP_OUTCOME)
    await render_booking_detail(callback, booking_id)
    await callback.answer(OUTCOME_TEXT.get(outcome, "Готово"))



@router.callback_query(F.data.startswith("a:bc:"))
async def booking_cancel_confirm(callback: CallbackQuery):
    booking_id = int(callback.data.split(":")[2])
    row = await db.get_booking(booking_id)
    if not row:
        await callback.answer("Заявка не найдена.", show_alert=True)
        return

    full_name, date_title = row[2], row[8]

    await show(
        callback,
        f"❌ Отменить заявку №{booking_id}?\n\n"
        f"Клиент: {full_name or '—'}\n"
        f"Дата: {date_title or '—'}\n\n"
        "Место освободится, клиент получит уведомление.",
        InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(
                text="❌ Да, отменить", callback_data=f"a:bc1:{booking_id}"
            )],
            [back_button(f"a:b:{booking_id}", "⬅️ Отмена")],
        ]),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("a:bc1:"))
async def booking_cancel_do(callback: CallbackQuery, bot: Bot):
    booking_id = int(callback.data.split(":")[2])
    changed, user_id = await db.cancel_booking(
        booking_id, callback.from_user.id, callback.from_user.full_name
    )

    if user_id is None:
        await callback.answer("Заявка не найдена.", show_alert=True)
        return

    if not changed:
        await callback.answer("Заявка уже была отменена.", show_alert=True)
        await render_booking_detail(callback, booking_id)
        return

    await callback.answer("Запись отменена")

    # Язык клиента, а не админа, который нажал кнопку.
    client_lang = texts.lang_or_default(await db.get_user_lang(user_id))
    try:
        await bot.send_message(user_id, texts.t("booking_cancelled", client_lang))
    except Exception as e:
        print(f"[client] не удалось уведомить {user_id} об отмене: {e}")
        await callback.message.answer(f"⚠️ Клиенту не доставлено уведомление: {e}")

    await render_booking_detail(callback, booking_id)


# ---------- СТАТИСТИКА ----------

@router.callback_query(F.data == "a:stats")
async def stats(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    data = await db.get_stats()
    dates = await db.get_dates_overview()

    lines = [
        "📊 Статистика",
        "",
        f"Дат всего: {data['dates_total']} (активных: {data['dates_active']})",
        f"Заявок всего: {data['total']}",
        f"✅ Подтверждено: {data['confirmed']}",
        f"⏳ Ждут подтверждения: {data['pending']}",
        f"❌ Отменено: {data['cancelled']}",
    ]

    ex = await db.get_outcome_stats()
    if ex['total']:
        lines += [
            "",
            f"🎓 Прошедшие экзамены: {ex['total']}",
            f"Явилось: {ex['came']}   ·   сдало: {ex['passed']}   ·   не сдало: {ex['failed']}",
            f"Не пришло: {ex['no_show']}   ·   неявки: {ex['no_show_pct']}% от проставленных",
        ]
        if ex['unmarked']:
            lines.append(f"Без итога: {ex['unmarked']} — их стоит проставить")

    if dates:
        lines += ["", "По датам:"]
        for _, title, seats_limit, is_active, confirmed, pending, _ in dates:
            taken = confirmed + pending
            seats = f"{taken}/{seats_limit}" if seats_limit else str(taken)
            mark = "" if is_active else " (скрыта)"
            lines.append(f"• {title}{mark}: {seats}  ✅{confirmed} ⏳{pending}")

    await show(callback, "\n".join(lines), InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📥 Экспорт в Excel", callback_data="a:export")],
        [back_button("a:menu", "⬅️ В меню")],
    ]))
    await callback.answer()


# ---------- ЭКСПОРТ ----------

async def build_workbook() -> bytes:
    rows = await db.get_export_rows()
    dates = await db.get_dates_overview()

    wb = Workbook()

    ws = wb.active
    ws.title = "Записи"
    headers = ["№", "Дата экзамена", "Кого записали", "Аккаунт", "Username",
               "Telegram ID", "Статус", "Итог", "Подтвердил", "Создана (Ташкент)"]
    ws.append(headers)
    for (booking_id, date_title, full_name, username, user_id, status,
         confirmed_by, created_at, applicant_name, outcome) in rows:
        ws.append([
            booking_id,
            date_title or "—",
            # с одного аккаунта могут записать друга — тогда это разные люди
            applicant_name or full_name or "—",
            full_name or "—",
            f"@{username}" if username else "—",
            user_id,
            STATUS_TEXT.get(status, status),
            OUTCOME_TEXT.get(outcome, "—"),
            confirmed_by or "—",
            fmt_dt(created_at),
        ])

    ws2 = wb.create_sheet("По датам")
    headers2 = ["Дата", "Лимит мест", "Подтверждено", "Ждут подтверждения",
                "Занято", "Свободно", "Активна"]
    ws2.append(headers2)
    for _, title, seats_limit, is_active, confirmed, pending, _ in dates:
        taken = confirmed + pending
        ws2.append([
            title,
            seats_limit or "без лимита",
            confirmed,
            pending,
            taken,
            max(seats_limit - taken, 0) if seats_limit else "—",
            "да" if is_active else "нет",
        ])

    for sheet, header_row in ((ws, headers), (ws2, headers2)):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        for i, name in enumerate(header_row, start=1):
            width = max(len(name) + 2, 12)
            for cell in sheet[get_column_letter(i)][1:]:
                width = max(width, min(len(str(cell.value or "")) + 2, 40))
            sheet.column_dimensions[get_column_letter(i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


PERIOD_NAME = {"today": "сегодня", "week": "неделя", "month": "месяц"}


@router.callback_query(F.data.startswith("a:fn:"))
async def funnel(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    period = callback.data.split(":")[2]
    if period not in PERIOD_NAME:
        period = "today"

    steps = await db.get_funnel(period)
    out = [f"🔻 Воронка · {PERIOD_NAME[period]}", ""]

    counts = {step: count for step, _, count in steps}
    started = counts.get(db.STEP_START, 0)
    prev = None
    for step, label, count in steps:
        # процент от предыдущего шага показывает, где именно отваливаются
        if prev is None:
            share = "—"
        elif prev == 0:
            share = "0%"
        else:
            share = f"{round(count * 100 / prev)}%"
        hint = "   ⏳" if step == db.STEP_OUTCOME else ""
        out.append(f"{label}: {count}   ({share} от пред.){hint}")
        prev = count

    if started:
        paid = counts.get(db.STEP_PAID, 0)
        out += [
            "",
            f"Конверсия в оплату: {round(paid * 100 / started)}% от запустивших",
            "",
            "⏳ результат проставляется через недели, после самого экзамена —",
            "в коротком периоде он почти всегда нулевой.",
        ]
    else:
        out += ["", "За этот период событий нет."]

    buttons = [InlineKeyboardButton(
        text=("· " + name + " ·") if key == period else name,
        callback_data=f"a:fn:{key}",
    ) for key, name in PERIOD_NAME.items()]

    await show(callback, "\n".join(out), InlineKeyboardMarkup(inline_keyboard=[
        buttons,
        [back_button("a:menu", "⬅️ В меню")],
    ]))
    await callback.answer()



@router.callback_query(F.data == "a:export")
async def export_excel(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.answer("Готовлю файл…")

    payload = await build_workbook()
    filename = f"zapisi_{datetime.now(config.TZ):%Y-%m-%d_%H-%M}.xlsx"

    await callback.message.answer_document(
        BufferedInputFile(payload, filename=filename),
        caption="📥 Выгрузка записей",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [back_button("a:menu", "⬅️ В меню")]
        ]),
    )
